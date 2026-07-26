import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from automation.utils.logger import logger

class ExcelReportGenerator:
    def __init__(self, test_results: list, summary_metrics: dict, output_dir: str = None):
        self.test_results = test_results
        self.summary_metrics = summary_metrics
        # e:\Medicine\automation\utils\excel_report_generator.py -> root_dir = e:\Medicine
        root_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        if not output_dir:
            self.output_dirs = [
                os.path.join(root_dir, 'Test Results', 'Excel'),
                os.path.join(root_dir, 'automation', 'reports', 'Excel')
            ]
        else:
            self.output_dirs = [output_dir]
            
        for d in self.output_dirs:
            os.makedirs(d, exist_ok=True)

    def generate_all_reports(self):
        logger.info("Generating Excel reports using openpyxl...")
        for d in self.output_dirs:
            self.generate_main_automation_report(d)
            self.generate_passed_tests_report(d)
            self.generate_failed_tests_report(d)
            self.generate_summary_report(d)
        logger.info(f"Excel reports successfully generated in {self.output_dirs}.")

    def _apply_header_style(self, cell):
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _apply_row_style(self, row, is_even=False, status="PASSED"):
        fill_color = 'F2F2F2' if is_even else 'FFFFFF'
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        for cell in row:
            cell.font = Font(name='Calibri', size=10)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    def _auto_fit_columns(self, ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

    def generate_main_automation_report(self, target_dir: str):
        filepath = os.path.join(target_dir, 'Automation_Test_Report.xlsx')
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        headers = ['Test ID', 'Module', 'Test Name', 'Status', 'Execution Time (s)', 'Priority', 'Failure Reason']

        # Sheet 1: Executed Test Cases
        ws1 = wb.create_sheet(title='Executed Test Cases')
        ws1.append(headers)
        for cell in ws1[1]:
            self._apply_header_style(cell)
        
        for idx, res in enumerate(self.test_results, start=2):
            row_data = [
                res.get('test_id', ''),
                res.get('module', ''),
                res.get('test_name', ''),
                res.get('status', ''),
                round(res.get('duration', 0.0), 3),
                res.get('priority', 'P1'),
                res.get('failure_reason', '')
            ]
            ws1.append(row_data)
            row_cells = ws1[idx]
            self._apply_row_style(row_cells, is_even=(idx % 2 == 0), status=res.get('status'))
            
            status_cell = row_cells[3]
            st = res.get('status', 'PASSED')
            if st == 'PASSED':
                status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                status_cell.font = Font(color='006100', bold=True)
            elif st == 'FAILED':
                status_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                status_cell.font = Font(color='9C0006', bold=True)
            elif st == 'SKIPPED':
                status_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                status_cell.font = Font(color='9C6500', bold=True)
        self._auto_fit_columns(ws1)

        # Sheet 2: Passed Tests
        ws2 = wb.create_sheet(title='Passed Tests')
        ws2.append(headers[:6])
        for cell in ws2[1]:
            self._apply_header_style(cell)
        passed_tests = [r for r in self.test_results if r.get('status') == 'PASSED']
        for idx, res in enumerate(passed_tests, start=2):
            row_data = [res.get('test_id'), res.get('module'), res.get('test_name'), 'PASSED', round(res.get('duration', 0.0), 3), res.get('priority')]
            ws2.append(row_data)
            self._apply_row_style(ws2[idx], is_even=(idx % 2 == 0), status='PASSED')
        self._auto_fit_columns(ws2)

        # Sheet 3: Failed Tests
        ws3 = wb.create_sheet(title='Failed Tests')
        ws3.append(headers)
        for cell in ws3[1]:
            self._apply_header_style(cell)
        failed_tests = [r for r in self.test_results if r.get('status') == 'FAILED']
        for idx, res in enumerate(failed_tests, start=2):
            row_data = [res.get('test_id'), res.get('module'), res.get('test_name'), 'FAILED', round(res.get('duration', 0.0), 3), res.get('priority'), res.get('failure_reason')]
            ws3.append(row_data)
            self._apply_row_style(ws3[idx], is_even=(idx % 2 == 0), status='FAILED')
        self._auto_fit_columns(ws3)

        # Sheet 4: Skipped Tests
        ws4 = wb.create_sheet(title='Skipped Tests')
        ws4.append(headers)
        for cell in ws4[1]:
            self._apply_header_style(cell)
        skipped_tests = [r for r in self.test_results if r.get('status') == 'SKIPPED']
        for idx, res in enumerate(skipped_tests, start=2):
            row_data = [res.get('test_id'), res.get('module'), res.get('test_name'), 'SKIPPED', round(res.get('duration', 0.0), 3), res.get('priority'), res.get('failure_reason', '')]
            ws4.append(row_data)
            self._apply_row_style(ws4[idx], is_even=(idx % 2 == 0), status='SKIPPED')
        self._auto_fit_columns(ws4)

        # Sheet 5: Execution Metrics
        ws5 = wb.create_sheet(title='Execution Metrics')
        ws5.append(['Metric Name', 'Value'])
        for cell in ws5[1]:
            self._apply_header_style(cell)
        metrics = [
            ('Total Executed Tests', self.summary_metrics.get('total', 0)),
            ('Passed Tests', self.summary_metrics.get('passed', 0)),
            ('Failed Tests', self.summary_metrics.get('failed', 0)),
            ('Skipped Tests', self.summary_metrics.get('skipped', 0)),
            ('Pass Percentage', f"{self.summary_metrics.get('pass_rate', 0.0):.2f}%"),
            ('Total Execution Duration (s)', f"{self.summary_metrics.get('duration', 0.0):.2f}s"),
            ('Environment URL', self.summary_metrics.get('base_url', '')),
            ('Execution Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ]
        for idx, (m_name, m_val) in enumerate(metrics, start=2):
            ws5.append([m_name, m_val])
            self._apply_row_style(ws5[idx], is_even=(idx % 2 == 0))
        self._auto_fit_columns(ws5)

        # Sheet 6: Defect Summary
        ws6 = wb.create_sheet(title='Defect Summary')
        ws6.append(['Defect ID', 'Module', 'Associated Test ID', 'Defect Title', 'Severity', 'Failure Traceback'])
        for cell in ws6[1]:
            self._apply_header_style(cell)
        for idx, res in enumerate(failed_tests, start=2):
            ws6.append([
                f"DEF-{idx-1:03d}",
                res.get('module'),
                res.get('test_id'),
                f"Failure in {res.get('test_name')}",
                res.get('priority', 'High'),
                res.get('failure_reason', 'Assertion / Element click error')
            ])
            self._apply_row_style(ws6[idx], is_even=(idx % 2 == 0), status='FAILED')
        self._auto_fit_columns(ws6)

        wb.save(filepath)

    def generate_passed_tests_report(self, target_dir: str):
        filepath = os.path.join(target_dir, 'Passed_Test_Cases.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Passed Test Cases"
        headers = ['Test ID', 'Module', 'Test Name', 'Status', 'Execution Time (s)', 'Priority']
        ws.append(headers)
        for cell in ws[1]:
            self._apply_header_style(cell)
        passed_tests = [r for r in self.test_results if r.get('status') == 'PASSED']
        for idx, res in enumerate(passed_tests, start=2):
            ws.append([res.get('test_id'), res.get('module'), res.get('test_name'), 'PASSED', round(res.get('duration', 0.0), 3), res.get('priority')])
            self._apply_row_style(ws[idx], is_even=(idx % 2 == 0), status='PASSED')
        self._auto_fit_columns(ws)
        wb.save(filepath)

    def generate_failed_tests_report(self, target_dir: str):
        filepath = os.path.join(target_dir, 'Failed_Test_Cases.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Failed Test Cases"
        headers = ['Test ID', 'Module', 'Test Name', 'Status', 'Execution Time (s)', 'Priority', 'Failure Reason']
        ws.append(headers)
        for cell in ws[1]:
            self._apply_header_style(cell)
        failed_tests = [r for r in self.test_results if r.get('status') == 'FAILED']
        for idx, res in enumerate(failed_tests, start=2):
            ws.append([res.get('test_id'), res.get('module'), res.get('test_name'), 'FAILED', round(res.get('duration', 0.0), 3), res.get('priority'), res.get('failure_reason')])
            self._apply_row_style(ws[idx], is_even=(idx % 2 == 0), status='FAILED')
        self._auto_fit_columns(ws)
        wb.save(filepath)

    def generate_summary_report(self, target_dir: str):
        filepath = os.path.join(target_dir, 'Summary_Report.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary Report"
        headers = ['Module', 'Total Cases', 'Passed', 'Failed', 'Skipped', 'Pass Rate (%)']
        ws.append(headers)
        for cell in ws[1]:
            self._apply_header_style(cell)

        module_stats = {}
        for res in self.test_results:
            mod = res.get('module', 'General')
            if mod not in module_stats:
                module_stats[mod] = {'total': 0, 'passed': 0, 'failed': 0, 'skipped': 0}
            module_stats[mod]['total'] += 1
            st = res.get('status', 'PASSED').lower()
            if st in module_stats[mod]:
                module_stats[mod][st] += 1

        for idx, (mod, stats) in enumerate(module_stats.items(), start=2):
            tot = stats['total']
            pas = stats['passed']
            fai = stats['failed']
            skp = stats['skipped']
            rate = (pas / tot * 100.0) if tot > 0 else 0.0
            ws.append([mod, tot, pas, fai, skp, f"{rate:.1f}%"])
            self._apply_row_style(ws[idx], is_even=(idx % 2 == 0))
        self._auto_fit_columns(ws)
        wb.save(filepath)
